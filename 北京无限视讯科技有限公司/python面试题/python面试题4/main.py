from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def get_data():
    """
    获取员工和水果信息，并保存到Excel文件中
    """
    driver = webdriver.Chrome()
    driver.get("http://ser988748769259.gptapp.ren/")

    driver.implicitly_wait(5)
    driver.find_element(by=By.XPATH, value='/html/body/div[1]/button[1]').click()  # 点击员工按钮
    # 等待first_iframe加载完成
    WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "firstIframe")))
    first_iframe = driver.find_element(by=By.ID, value="firstIframe")
    # 切换到id为firstIframe的iframe中
    driver.switch_to.frame(first_iframe)
    # 切换到子iframe中
    driver.switch_to.frame(driver.find_element(by=By.TAG_NAME, value="iframe"))

    # 查找body中的三个div元素
    divs = driver.find_elements(by=By.TAG_NAME, value="div")
    worker_list = []
    for div in divs:
        name = div.find_elements(by=By.TAG_NAME, value="p")[0].text.split(": ")[1]
        phone = div.find_elements(by=By.TAG_NAME, value="p")[1].text.split(": ")[1]
        sex = div.find_elements(by=By.TAG_NAME, value="p")[2].text.split(": ")[1]
        # 后续写入excel文件中
        worker_list.append({
            "name": name,
            "phone": phone,
            "sex": sex
        })

    driver.switch_to.default_content()  # 切换回主文档,即跳出所有iframe

    driver.find_element(by=By.XPATH, value='/html/body/div[1]/button[2]').click()  # 点击水果按钮
    # 重复上面等待iframe加载完成的操作
    WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "firstIframe")))
    first_iframe = driver.find_element(by=By.ID, value="firstIframe")
    driver.switch_to.frame(first_iframe)
    driver.switch_to.frame(driver.find_element(by=By.TAG_NAME, value="iframe"))

    # 查找body中的三个div元素
    divs = driver.find_elements(by=By.TAG_NAME, value="div")
    fruit_list = []
    for div in divs:
        name = div.find_elements(by=By.TAG_NAME, value="p")[0].text.split(": ")[1]
        price = div.find_elements(by=By.TAG_NAME, value="p")[1].text.split(": ")[1]
        # 后续写入excel文件中
        fruit_list.append({
            "name": name,
            "price": price
        })

        # 构建员工 DataFrame
        workers_df = pd.DataFrame(worker_list)
        # 构建水果 DataFrame
        fruits_df = pd.DataFrame(fruit_list)

        # 保存 Excel 文件
        with pd.ExcelWriter("output.xlsx") as writer:
            workers_df.to_excel(writer, sheet_name="员工", index=False)
            fruits_df.to_excel(writer, sheet_name="水果", index=False)

    driver.quit()


def center_align_cells(file_path):
    """
    在保存 Excel 文件后调用该函数来居中对齐单元格
    """
    wb = load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    wb.save(file_path)


get_data()
center_align_cells("output.xlsx")
