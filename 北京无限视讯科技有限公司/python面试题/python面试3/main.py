import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# 定义标黄填充样式
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# 读取xlsx文件
df = pd.read_excel('面试3.xlsx')

# 创建一个新的Excel以保存带样式的数据
output_wb = Workbook()
output_ws = output_wb.active  # 激活sheet

# 将原始数据写入新的Excel文件（相当于复制了个副本）
for r in dataframe_to_rows(df, index=False, header=True):
    output_ws.append(r)

# 遍历所有单元格，根据条件进行黄色填充
for row in output_ws.iter_rows(min_row=2, values_only=False):  # min_rows=2跳过标题行，values_only=False返回单元格对象
    row_num = row[0].row  # 获取行号
    df_row_index = row_num - 2  # 减去标题行的偏移量

    # 填充身高在140cm到220cm之间的单元格
    if 140 < df.loc[df_row_index, '身高(cm)'] < 220:
        for cell in row:
            if cell.column == 3:  # 身高列
                cell.fill = yellow_fill

    # 填充体重在40kg到150kg之外的单元格
    if not (40 < df.loc[df_row_index, '体重(kg)'] < 150):
        for cell in row:
            if cell.column == 4:  # 体重列
                cell.fill = yellow_fill

    # 填充血型为“不详”的单元格
    if df.loc[df_row_index, '血型'] == '不详':
        for cell in row:
            if cell.column == 9:  # 血型列
                cell.fill = yellow_fill

    # 填充空单元格或值为“无”的单元格
    for cell in row:
        column_name = df.columns[cell.column - 1]  # 获取列名
        if pd.isna(df.loc[df_row_index, column_name]) or df.loc[df_row_index, column_name] == '无':
            cell.fill = yellow_fill

# 保存填充后的Excel文件
output_wb.save('yellow_fill.xlsx')